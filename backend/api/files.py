"""统一阅读器后端（0.1.4 批 6 决策 10）：查看与入库分家。

GET /api/docs/{doc_id}/view   格式调度的结构化查看数据（大表分页）
GET /api/files/{doc_id}       原件流（pdf/图片给 WebView 原生渲染）

调度：md/txt→text · csv/xlsx/xls→table（sheet 页签+分页 500 行）·
docx→html（mammoth，缺则纯文本）· pdf→原件流+文本模式双轨 ·
图片→原件流 · url 文档→档案 md。
"""
from __future__ import annotations

import csv as _csv
import sys
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config
from api.deps import get_db
from ingestion import archiver

router = APIRouter(prefix="/api", tags=["files"])

_MIME = {".pdf": "application/pdf", ".png": "image/png",
         ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".gif": "image/gif",
         ".webp": "image/webp", ".html": "text/html; charset=utf-8",
         ".htm": "text/html; charset=utf-8"}
_IMG = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
PAGE_ROWS = 500


def _original(doc_id: str) -> Path | None:
    for p in config.SOURCE_FILES_PATH.glob(f"{doc_id}.*"):
        return p
    return None


def _read_text_fallback(p: Path) -> str:
    """utf-8 → gbk → gb18030 回退（决策 8：政府站/老文件编码）。"""
    for enc in ("utf-8", "gbk", "gb18030"):
        try:
            return p.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return p.read_text(encoding="utf-8", errors="replace")


@router.post("/files/reveal")
def reveal_dir(kind: str = "skills"):
    """0.1.8 Q7：用系统资源管理器打开数据目录（白名单限定，防任意路径）。"""
    import os
    dirs = {"skills": config.KNOWLEDGE_BASE_PATH / "skills",
            "fonts": config.KNOWLEDGE_BASE_PATH / "fonts",
            "root": config.KNOWLEDGE_BASE_PATH}
    p = dirs.get(kind)
    if p is None:
        raise HTTPException(422, f"未知目录 {kind}")
    p.mkdir(parents=True, exist_ok=True)
    os.startfile(str(p))  # noqa: S606 本地桌面软件，仅 Windows
    return {"opened": str(p)}


@router.get("/files/{doc_id}")
def get_file(doc_id: str):
    p = _original(doc_id)
    if p is None:
        raise HTTPException(404, f"原件不存在: {doc_id}")
    return FileResponse(p, media_type=_MIME.get(p.suffix.lower(),
                                                "application/octet-stream"),
                        filename=p.name)


def _open_text_stream(p: Path):
    """0.1.5 D3：流式打开——先用 64KB 样本定编码，再逐行读不整表进内存。"""
    for enc in ("utf-8", "gbk", "gb18030"):
        try:
            with open(p, "r", encoding=enc) as f:
                f.read(65536)
            return open(p, "r", encoding=enc, newline="")
        except UnicodeDecodeError:
            continue
    return open(p, "r", encoding="utf-8", errors="replace", newline="")


def _table_view(p: Path, sheet: str, page: int) -> dict:
    """0.1.5 D3：窗口只读——csv 逐行流式跳读；xlsx read_only 按 min/max_row 取页，
    total 用 max_row / 行计数，10 万行级内存平稳。"""
    ext = p.suffix.lower()
    start = max(page, 0) * PAGE_ROWS
    if ext == ".csv":
        rows: list[list[str]] = []
        total = 0
        with _open_text_stream(p) as f:
            for i, r in enumerate(_csv.reader(f)):
                if start <= i < start + PAGE_ROWS:
                    rows.append([str(c) for c in r])
                total = i + 1
        return {"kind": "table", "sheets": ["Sheet1"], "sheet": "Sheet1",
                "rows": rows, "total_rows": total,
                "page": page, "page_size": PAGE_ROWS}
    if ext == ".xls":
        # 0.1.5 D7：老格式走 xlrd（可选依赖）；缺依赖显式提示另存 .xlsx
        try:
            import xlrd
        except ImportError:
            raise HTTPException(501, "查看 .xls 需要 xlrd 组件（未安装）；"
                                     "可用「打开原件」或另存为 .xlsx")
        book = xlrd.open_workbook(str(p))
        sheets = book.sheet_names()
        cur = sheet if sheet in sheets else sheets[0]
        ws = book.sheet_by_name(cur)
        end = min(start + PAGE_ROWS, ws.nrows)
        rows = [["" if c in (None, "") else str(c) for c in ws.row_values(i)]
                for i in range(start, end)] if start < ws.nrows else []
        return {"kind": "table", "sheets": sheets, "sheet": cur,
                "rows": rows, "total_rows": ws.nrows,
                "page": page, "page_size": PAGE_ROWS}
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(501, "查看 Excel 需要 openpyxl（未安装）；"
                                 "可用「打开原件」交给系统程序")
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        cur = sheet if sheet in sheets else sheets[0]
        ws = wb[cur]
        rows = [["" if c is None else str(c) for c in r]
                for r in ws.iter_rows(min_row=start + 1,
                                      max_row=start + PAGE_ROWS,
                                      values_only=True)]
        total = ws.max_row or (start + len(rows))
    finally:
        wb.close()
    return {"kind": "table", "sheets": sheets, "sheet": cur,
            "rows": rows, "total_rows": total,
            "page": page, "page_size": PAGE_ROWS}


def _docx_view(p: Path) -> dict:
    try:
        import mammoth
        with open(p, "rb") as f:
            html = mammoth.convert_to_html(f).value
        return {"kind": "html", "content": html}
    except ImportError:
        pass
    try:
        import docx
        text = "\n\n".join(para.text for para in docx.Document(str(p)).paragraphs)
        return {"kind": "text", "format": "txt", "content": text,
                "note": "mammoth 未装，降级纯文本（无版式）"}
    except ImportError:
        raise HTTPException(501, "查看 docx 需要 mammoth 或 python-docx（均未安装）")


@router.get("/docs/{doc_id}/view")
def view_doc(doc_id: str, sheet: str = "", page: int = 0):
    doc = get_db().get_document(doc_id)
    if doc is None:
        raise HTTPException(404, f"文档不存在: {doc_id}")
    p = _original(doc_id)
    fmt, conversion = archiver.classify_conversion(
        str(p) if p else (doc.get("source_url") or ""),
        doc.get("source_type") or ("url" if doc.get("source_url") else "file"),
        "x" * 1000)   # 此处只取格式名；conversion 以档案 frontmatter 为准
    head = {"doc_id": doc_id, "title": doc.get("title") or doc_id,
            "author": doc.get("author") or "", "format": fmt,
            "conversion": conversion,
            "has_file": p is not None}

    if p is None:
        # URL 文档 / 原件缺失：档案 md → 标准化 md 兜底
        for cand in archiver.archive_paths(doc_id):
            if cand.endswith(".md"):
                return {**head, "kind": "text", "format": "md",
                        "content": Path(cand).read_text(encoding="utf-8")}
        for cand in config.STANCES_PATH.glob(f"*/{doc_id}.md"):
            return {**head, "kind": "text", "format": "md",
                    "content": cand.read_text(encoding="utf-8")}
        raise HTTPException(404, "无原件也无档案 md，无法查看")

    ext = p.suffix.lower()
    if ext in {".md", ".markdown", ".txt"}:
        return {**head, "kind": "text",
                "format": "md" if ext != ".txt" else "txt",
                "content": _read_text_fallback(p)}
    if ext in {".csv", ".xlsx", ".xls"}:
        return {**head, **_table_view(p, sheet, page)}
    if ext == ".docx":
        return {**head, **_docx_view(p)}
    if ext == ".pdf":
        # 双轨：原件流（WebView 原生 PDF 渲染）+ 文本模式（切块拼接）
        rows = get_db().conn.execute(
            "SELECT text FROM chunks WHERE doc_id=? ORDER BY chunk_id",
            (doc_id,)).fetchall()
        return {**head, "kind": "pdf", "file_url": f"/api/files/{doc_id}",
                "text": "\n\n".join(r["text"] for r in rows)}
    if ext in _IMG:
        return {**head, "kind": "image", "file_url": f"/api/files/{doc_id}"}
    return {**head, "kind": "file", "file_url": f"/api/files/{doc_id}",
            "note": "该格式暂不支持内嵌查看，可下载原件"}
